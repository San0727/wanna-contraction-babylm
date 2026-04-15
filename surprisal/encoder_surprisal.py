from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer, AutoModelForMaskedLM

import math
import pandas as pd
import torch
import torch.nn.functional as F

# Model Card

model_name = ""

# Load Model and Tokenizer

model = AutoModelForMaskedLM.from_pretrained(model_name)
tokenizer = AutoTokenizer.from_pretrained(model_name)
model.eval()
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

def compute_masked_surprisal_multitoken(sentence, target_word):
    try:
        device = next(model.parameters()).device  # get model's device

        # Tokenize target word
        target_tokens = tokenizer.tokenize(target_word)
        target_token_ids = tokenizer.convert_tokens_to_ids(target_tokens)

        if not target_token_ids:
            print(f"[WARNING] Target '{target_word}' tokenized into nothing.")
            return None
        if tokenizer.unk_token_id in target_token_ids:
            print(f"[WARNING] UNK token in target: {target_word} → {target_tokens}")
            return None

        # Expand [MASK] in sentence
        num_masks = len(target_token_ids)
        mask_expansion = " ".join([tokenizer.mask_token] * num_masks)
        masked_sentence = sentence.replace("[MASK]", mask_expansion, 1)

        # Tokenize updated sentence
        inputs = tokenizer(masked_sentence, return_tensors="pt")
        input_ids = inputs["input_ids"][0].to(device)
        attention_mask = inputs["attention_mask"][0].to(device)

        # Confirm correct number of [MASK]s
        mask_indices = (input_ids == tokenizer.mask_token_id).nonzero(as_tuple=True)[0]
        if len(mask_indices) != num_masks:
            print(f"[WARNING] Mismatch: {masked_sentence} vs {target_tokens}")
            return None

        total_surprisal = 0
        current_input_ids = input_ids.clone()

        for i, token_id in enumerate(target_token_ids):
            mask_index = mask_indices[i].item()

            with torch.no_grad():
                outputs = model(
                    input_ids=current_input_ids.unsqueeze(0),
                    attention_mask=attention_mask.unsqueeze(0)
                )
                logits = outputs.logits[0, mask_index]
                probs = torch.nn.functional.softmax(logits, dim=-1)

            prob = probs[token_id].item()
            total_surprisal += -math.log2(prob)

            # Replace [MASK] with actual subtoken
            current_input_ids[mask_index] = token_id

        return round(total_surprisal, 4)

    except Exception as e:
        print(f"[ERROR] sentence='{sentence}', target='{target_word}' → {e}")
        return None

# Test Sentences

sentence_pairs = [
    ("Who do you wanna apologize after the conference [MASK]?", "tomorrow"),
    ("Who do you wanna congratulate after the conference [MASK]?", "tomorrow")
]

# Compute Surprisal

for sentence, item in sentence_pairs:
    print("\n" + "-" * 50)
    print(f"{'Sentence:':<10} {sentence}")
    print(f"{'Target:':<10} {item}")
    print("-" * 50)

    surprisal = compute_masked_surprisal_multitoken(sentence, item)
    print(f"{item:>10} : {surprisal:.5f}")

# Load Excel File

df = pd.read_excel(".xlsx")

# Add Surprisal Column

df["SURPRISAL"] = df.apply(lambda row: compute_masked_surprisal_multitoken(row["SENTENCE"], row["ITEM"]), axis = 1)

# Save Results

output_file = ".xlsx"

with pd.ExcelWriter(output_file, engine = "openpyxl") as writer:
    df.to_excel(writer, index = False, sheet_name = "Sheet1")
    ws = writer.book["Sheet1"]

    # Set Font
    font = Font(name = "Arial", size = 12)

    # Set Column Widths
    column_widths = {1: 20, 2: 70, 3: 15, 4: 15}
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Apply Font and Alignment
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            cell.font = font
            if row_idx == 0:
                # Header Row: Center Everything
                cell.alignment = Alignment(horizontal = "center", vertical = "center")
            else:
                if col_idx == 0:
                    cell.alignment = Alignment(horizontal = "center", vertical = "center")
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal = "left", vertical = "center")
                elif col_idx == 2:
                    cell.alignment = Alignment(horizontal = "center", vertical = "center")
                elif col_idx == 3:
                    cell.alignment = Alignment(horizontal = "right", vertical = "center")
